# -*- coding: utf-8 -*-

from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active
ws.title = "New Title"
ws1 = wb.create_sheet("Mysheet1") 
ws2 = wb.create_sheet("Mysheet2") 
ws3 = wb.create_sheet("Mysheet3") 
#ws['A1'] = 42
#ws.append([1,2,3])
#ws['A2'] = datetime.datetime.now()
for i in range(1,16):
	for j in range(i,16):
		if(j<9):
			print 'Mysheet '+str(i)+' '+str(j)
			ws1.cell(column=j, row=i, value=i+j)
		elif(j>8 and j<12):
			print 'Mysheet1'
			print i,j,j-8
			ws2.cell(column=j-8, row=i, value=j+i)
		elif(j>11 and j<16):
			print 'Mysheet2'
			print i,j,j-11
			ws3.cell(column=j-11, row=i, value=j+i)

wb.save("sample1.xlsx")