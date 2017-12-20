# -*- coding: utf-8 -*-

#import xlwt
from openpyxl import Workbook
import datetime
from libtbx import easy_pickle
from scitbx.array_family import flex

begin = datetime.datetime.now()
data = easy_pickle.load("~/Desktop/pisadb/pisaDB.nl")
length = len(data)
data_type = type(data)
diff_list = [[0 for col in range(length)] for row in range(length)]
count = 0

#wb = xlwt.Workbook()
#ws = wb.add_sheet('part of the diff_matrix')

wb = Workbook()
ws = wb.active
ws.title = "dirr_col_250"
ws1 = wb.create_sheet("dirr_col_500") 
ws2 = wb.create_sheet("dirr_col_750") 
ws3 = wb.create_sheet("dirr_col_1000")

for i in range(1, 1001):
	for j in range(i, 1001):
		diff_list[i][j] = data[i] - data[j]
		if(j<251):
			ws.cell(column=j, row=i, value=diff_list[i][j].norm())
		elif(j>250 and j<501):
			ws1.cell(column=j-250, row=i, value=diff_list[i][j].norm())
		elif(j>500 and j<751):
			ws2.cell(column=j-500, row=i, value=diff_list[i][j].norm())
		else:
			ws3.cell(column=j-750, row=i, value=diff_list[i][j].norm())
		if(j%9==0):
			print count
		count = count + 1

wb.save('diff1000.xlsx')
end = datetime.datetime.now()
#print data
print str(end-begin)
print (data[0] - data[1]).norm()
print length
